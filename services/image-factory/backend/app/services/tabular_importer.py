"""Tabular import adapters (plan section 8: `tabular_importer`).

Reads a CSV or Excel (.xlsx/.xlsm) upload and returns a single internal schema:
(fieldnames, rows) where each row is a dict keyed by the header names.

Both adapters normalise headers to lower-case/stripped so downstream validation
is source-agnostic.
"""

from __future__ import annotations

import csv
import io

from app.core.config import settings

CSV_EXTS = (".csv", ".txt")
EXCEL_EXTS = (".xlsx", ".xlsm", ".xltx", ".xltm")


class TabularError(ValueError):
    pass


def _decode(raw: bytes) -> str:
    for encoding in ("utf-8-sig", "utf-8", "cp1258", "latin-1"):
        try:
            return raw.decode(encoding)
        except UnicodeDecodeError:
            continue
    return raw.decode("utf-8", errors="replace")


def _norm(value) -> str:
    return str(value).strip() if value is not None else ""


def _read_csv(raw: bytes) -> tuple[list[str], list[dict]]:
    text = _decode(raw)
    reader = csv.DictReader(io.StringIO(text))
    fieldnames = [(_norm(h)).lower() for h in (reader.fieldnames or [])]
    rows: list[dict] = []
    for raw_row in reader:
        row = {}
        for key, val in raw_row.items():
            if key is None:
                continue
            row[_norm(key).lower()] = _norm(val)
        rows.append(row)
    return fieldnames, rows


def _read_excel(raw: bytes) -> tuple[list[str], list[dict]]:
    from openpyxl import load_workbook

    try:
        wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    except Exception as exc:  # noqa: BLE001 - surface a friendly message
        raise TabularError(f"Không đọc được file Excel: {exc}") from exc

    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        raise TabularError("File Excel rỗng") from None

    fieldnames = [_norm(h).lower() for h in header_row if _norm(h)]
    # Map column index -> header (skip empty header columns).
    col_map = {i: _norm(h).lower() for i, h in enumerate(header_row) if _norm(h)}

    rows: list[dict] = []
    for excel_row in rows_iter:
        if excel_row is None:
            continue
        row = {col_map[i]: _norm(v) for i, v in enumerate(excel_row) if i in col_map}
        # Skip fully-empty rows.
        if any(row.values()):
            rows.append(row)
    wb.close()
    return fieldnames, rows


def read_rows(filename: str | None, raw: bytes) -> tuple[list[str], list[dict]]:
    """Return (fieldnames, rows) from a CSV or Excel upload."""
    if len(raw) > settings.max_csv_bytes:
        raise TabularError(f"File vượt quá kích thước tối đa {settings.max_csv_bytes} bytes")

    name = (filename or "").lower()
    if name.endswith(EXCEL_EXTS):
        return _read_excel(raw)
    if name.endswith(CSV_EXTS) or not name:
        return _read_csv(raw)
    # Unknown extension: sniff — Excel files are ZIP archives starting with "PK".
    if raw[:2] == b"PK":
        return _read_excel(raw)
    return _read_csv(raw)
